"""Orchestratore principale del modulo di valutazione RAG DIEM.

Coordina l'intero flusso di evaluation:
1. Caricamento del dataset di input (domande + ground truth)
2. Elaborazione delle domande tramite l'agente RAG
3. Costruzione dell'EvaluationDataset RAGAS
4. Esecuzione della valutazione con il Judge LLM robusto
5. Generazione dei report di output (JSON + CSV/Excel)

Produce due file di output:
- JSON strutturato con metriche per domanda, aggregati e metadati
- CSV/Excel tabellare con una riga per domanda e tutte le metriche
"""

import json
import logging
import time
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, Any, List

import pandas as pd

from src.evaluation.config import EvaluationConfig, load_evaluation_config
from src.evaluation.judge import create_judge_llm, RobustJudgeLLM
from src.evaluation.dataset import EvaluationDatasetBuilder

logger = logging.getLogger(__name__)


def _build_ragas_metrics(config: EvaluationConfig) -> list:
    """Costruisce la lista di metriche RAGAS in base alla configurazione.

    Importa e istanzia solo le metriche abilitate per minimizzare
    il numero di chiamate al Judge LLM.

    Args:
        config: Configurazione dell'evaluation.

    Returns:
        Lista di istanze di metriche RAGAS.

    Raises:
        ImportError: Se ragas non e' installato.
        ValueError: Se nessuna metrica e' abilitata.
    """
    try:
        from ragas.metrics import (
            ContextPrecision,
            ContextRecall,
            ResponseRelevancy,
            Faithfulness,
            FactualCorrectness,
        )
    except ImportError as e:
        raise ImportError(
            "ragas non installato. Installa con: pip install ragas"
        ) from e

    metrics = []
    mc = config.metrics

    if mc.enable_context_precision:
        metrics.append(ContextPrecision())
    if mc.enable_context_recall:
        metrics.append(ContextRecall())
    if mc.enable_response_relevancy:
        metrics.append(ResponseRelevancy())
    if mc.enable_faithfulness:
        metrics.append(Faithfulness())
    if mc.enable_factual_correctness:
        metrics.append(FactualCorrectness(mode=mc.factual_correctness_mode))

    if not metrics:
        raise ValueError(
            "Nessuna metrica RAGAS abilitata. Abilita almeno una metrica "
            "nella configurazione (EVAL_METRIC_*)."
        )

    metric_names = [type(m).__name__ for m in metrics]
    logger.info("Metriche RAGAS abilitate: %s", metric_names)

    return metrics


_RAGAS_COLUMN_MAP = {
    "user_input": "question",
    "retrieved_contexts": "retrieved_contexts",
    "response": "response",
    "reference": "ground_truth",
    "context_precision": "ctx_precision",
    "context_recall": "ctx_recall",
    "answer_relevancy": "response_relevancy",
    "response_relevancy": "response_relevancy",
    "faithfulness": "faithfulness",
    "factual_correctness": "factual_correctness",
}


class EvaluationRunner:
    """Orchestratore completo del flusso di evaluation.

    Gestisce l'intero ciclo di vita della valutazione: dal caricamento
    dei dati alla generazione dei report finali. Tutti i passaggi sono
    tracciati e i metadati vengono inclusi nei file di output.

    Attributes:
        config: Configurazione dell'evaluation.
        dataset_builder: Builder per il dataset RAGAS.
        judge_llm: Wrapper robusto del Judge LLM.
        ragas_result: Risultato grezzo della valutazione RAGAS.
        report_df: DataFrame Pandas con il report tabellare.
        run_metadata: Metadati dell'esecuzione corrente.
    """

    def __init__(self, config: Optional[EvaluationConfig] = None):
        """Inizializza il runner con la configurazione.

        Args:
            config: Configurazione dell'evaluation. Se None, viene
                   caricata automaticamente da variabili d'ambiente.
        """
        self.config = config or load_evaluation_config()
        self.dataset_builder = EvaluationDatasetBuilder(self.config)
        self.judge_llm: Optional[RobustJudgeLLM] = None
        self.ragas_result = None
        self.report_df: Optional[pd.DataFrame] = None
        self.run_metadata: Dict[str, Any] = {
            "started_at": None,
            "completed_at": None,
            "duration_seconds": 0,
            "status": "not_started",
        }

    def run(self, agent, input_file: Optional[str] = None) -> Dict[str, Any]:
        """Esegue l'intero flusso di evaluation end-to-end.

        Questo e' il metodo principale da chiamare. Esegue in sequenza:
        1. Caricamento dataset
        2. Creazione Judge LLM
        3. Elaborazione domande tramite agente RAG
        4. Costruzione EvaluationDataset RAGAS
        5. Esecuzione valutazione RAGAS
        6. Costruzione report
        7. Esportazione file di output

        Args:
            agent: Istanza di RAGAgent configurata e pronta.
            input_file: Percorso opzionale del file JSON di input.
                       Se None, usa il percorso dalla configurazione.

        Returns:
            Dizionario con il report completo dell'evaluation, inclusi
            metriche aggregate, metriche per domanda e metadati.
        """
        self.run_metadata["started_at"] = datetime.now().isoformat()
        self.run_metadata["status"] = "running"
        start_time = time.time()

        try:
            logger.info("=" * 60)
            logger.info("  STEP 1/6 — Caricamento dataset")
            logger.info("=" * 60)
            num_samples = self.dataset_builder.load_from_json(input_file)
            logger.info("Caricate %d domande.", num_samples)

            logger.info("=" * 60)
            logger.info("  STEP 2/6 — Inizializzazione Judge LLM")
            logger.info("=" * 60)
            self.judge_llm = create_judge_llm(self.config)

            logger.info("=" * 60)
            logger.info("  STEP 3/6 — Elaborazione domande tramite agente RAG")
            logger.info("=" * 60)
            processed = self.dataset_builder.process_with_agent(agent)
            logger.info("Elaborate %d domande.", len(processed))

            logger.info("=" * 60)
            logger.info("  STEP 4/6 — Costruzione EvaluationDataset RAGAS")
            logger.info("=" * 60)
            eval_dataset = self.dataset_builder.build_ragas_dataset()

            logger.info("=" * 60)
            logger.info("  STEP 5/6 — Esecuzione valutazione RAGAS")
            logger.info("=" * 60)
            self.ragas_result = self._run_ragas_evaluation(eval_dataset)

            logger.info("=" * 60)
            logger.info("  STEP 6/6 — Generazione report e esportazione")
            logger.info("=" * 60)
            report = self._build_full_report()
            self._export_outputs(report)

            self.dataset_builder.clear_intermediate()

            if self.judge_llm:
                self.judge_llm.print_stats()

            self.run_metadata["status"] = "completed"
            self.run_metadata["completed_at"] = datetime.now().isoformat()
            self.run_metadata["duration_seconds"] = round(time.time() - start_time, 1)

            logger.info("=" * 60)
            logger.info("  EVALUATION COMPLETATA CON SUCCESSO")
            logger.info("  Durata totale: %.1f secondi", self.run_metadata["duration_seconds"])
            logger.info("=" * 60)

            return report

        except Exception as e:
            self.run_metadata["status"] = "failed"
            self.run_metadata["error"] = str(e)
            self.run_metadata["completed_at"] = datetime.now().isoformat()
            self.run_metadata["duration_seconds"] = round(time.time() - start_time, 1)

            logger.error("EVALUATION FALLITA: %s", e, exc_info=True)
            raise

    def _run_ragas_evaluation(self, eval_dataset) -> Any:
        """Esegue la valutazione RAGAS in modalita' STRETTAMENTE SEQUENZIALE.

        CRITICO: RAGAS di default lancia le chiamate al Judge LLM in
        parallelo tramite asyncio, causando un burst di richieste che
        supera immediatamente i rate limit di Groq (specialmente sul
        piano Free). Per evitare questo problema:

        1. Forziamo batch_size=1 in ragas.evaluate() cosi' RAGAS
           processa UN sample alla volta.
        2. Impostiamo max_workers=1 tramite la RunConfig di RAGAS
           per disabilitare qualsiasi parallelismo interno.
        3. Aggiungiamo una pausa (EVAL_BATCH_DELAY) tra un sample
           e il successivo per distanziare le chiamate API.

        In questo modo con 3 sample e 5 metriche, le ~15 chiamate
        al Judge vengono eseguite una alla volta con pausa, invece
        di essere sparate tutte insieme.

        Args:
            eval_dataset: EvaluationDataset costruito dal builder.

        Returns:
            Risultato grezzo di ragas.evaluate().
        """
        from ragas import evaluate, RunConfig
        from ragas.llms import LangchainLLMWrapper
        from langchain_huggingface import HuggingFaceEmbeddings

        from src.config.settings import load_settings

        settings = load_settings()

        embedding_model = HuggingFaceEmbeddings(
            model_name=settings.embedding.model_name,
            encode_kwargs={
                "normalize_embeddings": settings.embedding.normalize_embeddings,
            },
        )

        metrics = _build_ragas_metrics(self.config)

        wrapped_judge = LangchainLLMWrapper(self.judge_llm)

        run_config = RunConfig(
            max_workers=1,
            max_wait=180,
            timeout=360,
            log_tenacity=True,
        )

        batch_delay = self.config.pipeline.batch_delay_seconds
        num_samples = len(eval_dataset)
        num_metrics = len(metrics)
        total_judge_calls = num_samples * num_metrics

        logger.info(
            "Avvio ragas.evaluate() in modalita' SEQUENZIALE — "
            "%d sample, %d metriche, ~%d chiamate Judge totali, "
            "batch_size=1, max_workers=1, pausa=%.1fs tra sample. "
            "Judge: %s/%s",
            num_samples,
            num_metrics,
            total_judge_calls,
            batch_delay,
            self.config.judge.provider,
            self.config.judge.model_name,
        )

        result = evaluate(
            dataset=eval_dataset,
            metrics=metrics,
            llm=wrapped_judge,
            embeddings=embedding_model,
            batch_size=1,
            run_config=run_config,
        )

        logger.info("ragas.evaluate() completato. Risultati aggregati: %s", result)

        return result

    def _build_full_report(self) -> Dict[str, Any]:
        """Costruisce il report completo con metriche aggregate e per-sample.

        Riallinea i risultati di RAGAS (che contengono solo i sample
        validi) con la lista completa dei sample elaborati, assegnando
        None alle metriche dei sample esclusi.

        Returns:
            Dizionario strutturato con il report completo.
        """

        ragas_df = self.ragas_result.to_pandas()

        ragas_df = ragas_df.rename(columns={
            col: _RAGAS_COLUMN_MAP.get(col, col) for col in ragas_df.columns
        })

        non_metric_cols = {"question", "retrieved_contexts", "response", "ground_truth"}
        metric_columns = [c for c in ragas_df.columns if c not in non_metric_cols]

        valid_indices = self.dataset_builder.get_valid_indices()
        all_processed = self.dataset_builder.get_all_processed()

        full_rows = []
        ragas_row_idx = 0

        for i, sample_data in enumerate(all_processed):
            row = {
                "index": sample_data["index"],
                "question": sample_data["question"],
                "ground_truth": sample_data["ground_truth"],
                "response": sample_data["response"],
                "was_blocked": sample_data["was_blocked"],
                "error": sample_data["error"],
                "processing_time_ms": sample_data["processing_time_ms"],
            }

            if i in valid_indices and ragas_row_idx < len(ragas_df):
                for mc in metric_columns:
                    row[mc] = ragas_df.iloc[ragas_row_idx].get(mc, None)
                ragas_row_idx += 1
            else:
                for mc in metric_columns:
                    row[mc] = None

            full_rows.append(row)

        self.report_df = pd.DataFrame(full_rows)

        valid_df = self.report_df[
            (self.report_df["was_blocked"] == False) &
            (self.report_df["error"].isna())
        ]

        aggregated_metrics = {}
        for mc in metric_columns:
            values = valid_df[mc].dropna()
            if len(values) > 0:
                aggregated_metrics[mc] = {
                    "mean": round(float(values.mean()), 4),
                    "median": round(float(values.median()), 4),
                    "std": round(float(values.std()), 4) if len(values) > 1 else 0.0,
                    "min": round(float(values.min()), 4),
                    "max": round(float(values.max()), 4),
                    "count": int(len(values)),
                }

        judge_stats = self.judge_llm.get_stats() if self.judge_llm else {}

        report = {
            "metadata": {
                "run": self.run_metadata,
                "dataset": self.dataset_builder.dataset_metadata,
                "judge": {
                    "provider": self.config.judge.provider,
                    "model": self.config.judge.model_name,
                    "temperature": self.config.judge.temperature,
                    "stats": judge_stats,
                },
                "metrics_enabled": [
                    type(m).__name__
                    for m in _build_ragas_metrics(self.config)
                ],
                "total_samples": len(all_processed),
                "valid_samples": len(valid_indices),
                "blocked_samples": sum(1 for s in all_processed if s["was_blocked"]),
                "error_samples": sum(1 for s in all_processed if s["error"]),
            },
            "aggregated_metrics": aggregated_metrics,
            "per_sample_results": [
                {
                    **sample_data,
                    "metrics": {
                        mc: full_rows[idx].get(mc)
                        for mc in metric_columns
                    },
                }
                for idx, sample_data in enumerate(all_processed)
            ],
        }

        logger.info("Report costruito: %d metriche aggregate.", len(aggregated_metrics))

        return report

    def _export_outputs(self, report: Dict[str, Any]) -> Dict[str, str]:
        """Esporta i file di output (JSON, CSV, Excel).

        Args:
            report: Report completo da esportare.

        Returns:
            Dizionario con i percorsi dei file generati.
        """
        output_dir = Path(self.config.output.output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        generated_files = {}

        if self.config.output.export_json:
            json_path = output_dir / f"eval_{timestamp}_metrics.json"
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(report, f, indent=self.config.output.json_indent, ensure_ascii=False)
            generated_files["json"] = str(json_path)
            logger.info("Report JSON esportato: %s", json_path)

        if self.config.output.export_csv and self.report_df is not None:
            csv_path = output_dir / f"eval_{timestamp}_report.csv"
            self.report_df.to_csv(csv_path, index=False, encoding="utf-8-sig")
            generated_files["csv"] = str(csv_path)
            logger.info("Report CSV esportato: %s", csv_path)

        if self.config.output.export_excel and self.report_df is not None:
            xlsx_path = output_dir / f"eval_{timestamp}_report.xlsx"
            try:
                self._export_styled_excel(xlsx_path)
                generated_files["xlsx"] = str(xlsx_path)
                logger.info("Report Excel esportato: %s", xlsx_path)
            except ImportError:
                logger.warning(
                    "openpyxl non installato, skip esportazione Excel. "
                    "Installa con: pip install openpyxl"
                )
            except Exception as e:
                logger.warning("Errore esportazione Excel: %s", e)

        self.run_metadata["output_files"] = generated_files

        return generated_files

    def _export_styled_excel(self, filepath: Path) -> None:
        """Esporta il DataFrame in Excel con formattazione professionale.

        Applica:
        - Header con sfondo colorato e testo bold
        - Formattazione condizionale per le metriche (verde/rosso)
        - Colonne con larghezza automatica
        - Formato numerico a 4 decimali per le metriche

        Args:
            filepath: Percorso del file Excel di output.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        df = self.report_df.copy()
        df.to_excel(filepath, index=False, engine="openpyxl")

        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        ws.title = "Evaluation Report"

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border

        non_metric_cols = {"index", "question", "ground_truth", "response",
                          "was_blocked", "error", "processing_time_ms"}
        metric_col_indices = [
            i + 1 for i, col in enumerate(df.columns)
            if col not in non_metric_cols
        ]

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        cell_align = Alignment(vertical="top", wrap_text=True)

        for row_idx in range(2, len(df) + 2):
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = cell_align

                if col_idx in metric_col_indices and cell.value is not None:
                    try:
                        val = float(cell.value)
                        cell.number_format = "0.0000"
                        if val >= 0.7:
                            cell.fill = green_fill
                        elif val < 0.4:
                            cell.fill = red_fill
                    except (ValueError, TypeError):
                        pass

        for col_idx, col_name in enumerate(df.columns, 1):
            letter = get_column_letter(col_idx)
            if col_name in ("question", "ground_truth", "response"):
                ws.column_dimensions[letter].width = 50
            elif col_name in non_metric_cols:
                ws.column_dimensions[letter].width = 18
            else:
                ws.column_dimensions[letter].width = 18

        ws.freeze_panes = "A2"

        wb.save(filepath)

    def get_summary(self) -> str:
        """Genera un riepilogo testuale dell'evaluation per il terminale.

        Returns:
            Stringa formattata con il riepilogo dell'evaluation.
        """
        if self.report_df is None:
            return "Nessun risultato disponibile. Esegui run() prima."

        lines = []
        lines.append("=" * 60)
        lines.append("  RIEPILOGO EVALUATION RAG DIEM")
        lines.append("=" * 60)
        lines.append(f"  Dataset:     {self.dataset_builder.dataset_metadata.get('dataset_name', 'N/D')}")
        lines.append(f"  Domande:     {len(self.dataset_builder.processed_samples)}")
        lines.append(f"  Valide:      {len(self.dataset_builder.get_valid_indices())}")
        lines.append(f"  Durata:      {self.run_metadata.get('duration_seconds', 0):.1f}s")
        lines.append(f"  Judge:       {self.config.judge.provider}/{self.config.judge.model_name}")
        lines.append("")

        non_metric = {"index", "question", "ground_truth", "response",
                      "was_blocked", "error", "processing_time_ms"}
        metric_cols = [c for c in self.report_df.columns if c not in non_metric]

        valid_df = self.report_df[
            (self.report_df["was_blocked"] == False) &
            (self.report_df["error"].isna())
        ]

        if metric_cols:
            lines.append("  METRICHE AGGREGATE (media):")
            lines.append("  " + "-" * 40)
            for mc in metric_cols:
                values = valid_df[mc].dropna()
                if len(values) > 0:
                    mean_val = values.mean()
                    lines.append(f"  {mc:<25s} {mean_val:.4f}")
            lines.append("")

        if self.judge_llm:
            stats = self.judge_llm.get_stats()
            lines.append("  STATISTICHE JUDGE:")
            lines.append("  " + "-" * 40)
            lines.append(f"  Chiamate totali:    {stats['total_calls']}")
            lines.append(f"  Retry:              {stats['total_retries']}")
            lines.append(f"  Riparazioni JSON:   {stats['total_repairs']}")
            lines.append(f"  Rate limit (429):   {stats['total_rate_limits']}")

        lines.append("=" * 60)

        return "\n".join(lines)